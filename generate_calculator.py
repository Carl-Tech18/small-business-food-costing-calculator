from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Food Cost Calculator"

header_font = Font(bold=True)
input_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
center_align = Alignment(horizontal="center")

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15

headers = ["Ingredient", "Unit", "Unit Cost", "Quantity Used", "Cost"]
ws.append(headers)
for col in range(1, 6):
    ws.cell(row=1, column=col).font = header_font
    ws.cell(row=1, column=col).alignment = center_align

ingredients = [
    ["Chicken", "kg", 150, 0.3],
    ["Onion", "pc", 5, 2],
    ["Garlic", "bulb", 8, 1],
    ["Oil", "liter", 120, 0.05],
    ["Rice", "kg", 50, 0.25]
]

for i, ingredient in enumerate(ingredients, start=2):
    for j, value in enumerate(ingredient, start=1):
        cell = ws.cell(row=i, column=j, value=value)
        if j > 2:
            cell.fill = input_fill
    ws.cell(row=i, column=5, value=f"=C{i}*D{i}").fill = total_fill

total_row = len(ingredients) + 2
ws.cell(row=total_row, column=4, value="TOTAL").font = header_font
ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})").fill = total_fill
ws.cell(row=total_row, column=5).font = header_font

summary_start = total_row + 2
summary_data = [
    ("Total Ingredient Cost", f"=E{total_row}"),
    ("Packaging Cost", 5),
    ("Other Cost (Gas, etc.)", 10),
    ("Total Cost", f"=B{summary_start+0}+B{summary_start+1}+B{summary_start+2}"),
    ("Desired Profit", 30),
    ("Selling Price", f"=B{summary_start+3}+B{summary_start+4}")
]

for i, (label, value) in enumerate(summary_data):
    label_cell = ws.cell(row=summary_start + i, column=1, value=label)
    value_cell = ws.cell(row=summary_start + i, column=2, value=value)
    label_cell.font = header_font
    if i not in [0, 3, 5]:
        value_cell.fill = input_fill
    else:
        value_cell.fill = total_fill
        value_cell.font = header_font

unit_dropdown = DataValidation(
    type="list",
    formula1='"kg,pc,bulb,liter,g"',
    showDropDown=True
)
ws.add_data_validation(unit_dropdown)
for row in range(2, 7):
    unit_dropdown.add(ws[f'B{row}'])

wb.save("Food_Costing_Calculator_v1.0.xlsx")
